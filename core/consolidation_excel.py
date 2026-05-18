"""
Consolidation Excel generator — builds a complete workbook from scratch
with Normalized P&L, EBITDA Bridge, Balance Sheet, and Adjustments Detail tabs.
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ── Styles ───────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SUBHEADER_FONT = Font(bold=True, size=10)
_TOTAL_FONT = Font(bold=True, size=10)
_TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
_NUM_FMT = '#,##0'
_NUM_FMT_DEC = '#,##0.00'
_PCT_FMT = '0.0%'


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _SUBHEADER_FONT
        cell.fill = _SUBHEADER_FILL


def _style_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL


def _write_num(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)):
        cell.number_format = _NUM_FMT
    return cell


# ── P&L line items (order matters) ───────────────────────────────────────────

_PNL_LINES = [
    ("sales_and_services", "Sales & Service Revenue", False),
    ("other_revenues", "Other Revenues", False),
    (None, "Total Revenue", True),
    ("cost_of_goods_sold", "Cost of Goods Sold", False),
    (None, "Gross Profit", True),
    ("selling_expenses", "Selling Expenses", False),
    ("administrative_expenses", "Administrative Expenses", False),
    ("other_expenses", "Other Expenses", False),
    (None, "EBITDA", True),
    ("depreciation_amortization", "Depreciation & Amortization", False),
    (None, "EBIT", True),
    ("interest_expense", "Interest Expense", False),
    (None, "Profit Before Tax", True),
    ("income_tax", "Income Tax", False),
    (None, "Net Profit", True),
]

_BS_LINES = [
    ("total_assets", "Total Assets"),
    ("total_liabilities", "Total Liabilities"),
    ("total_equity", "Total Equity"),
    ("cash", "Cash & Cash Equivalents"),
    ("total_debt", "Total Debt"),
    ("accounts_receivable", "Trade Receivables"),
    ("inventory", "Inventory"),
    ("accounts_payable", "Trade Payables"),
    ("related_party_loans", "Related-Party Loans"),
]


def _calc_derived(key, pnl_data, year):
    """Calculate derived P&L totals."""
    def _v(k):
        return (pnl_data.get(k) or {}).get(year) or 0

    if key == "Total Revenue":
        return _v("sales_and_services") + _v("other_revenues")
    elif key == "Gross Profit":
        return (_v("sales_and_services") + _v("other_revenues")) - abs(_v("cost_of_goods_sold"))
    elif key == "EBITDA":
        rev = _v("sales_and_services") + _v("other_revenues")
        gp = rev - abs(_v("cost_of_goods_sold"))
        return gp - abs(_v("selling_expenses")) - abs(_v("administrative_expenses")) - abs(_v("other_expenses"))
    elif key == "EBIT":
        ebitda = _calc_derived("EBITDA", pnl_data, year)
        return ebitda - abs(_v("depreciation_amortization"))
    elif key == "Profit Before Tax":
        ebit = _calc_derived("EBIT", pnl_data, year)
        return ebit - abs(_v("interest_expense"))
    elif key == "Net Profit":
        pbt = _calc_derived("Profit Before Tax", pnl_data, year)
        return pbt - abs(_v("income_tax"))
    return 0


def _apply_adjustments(pnl_data: dict, adjustments: list, entity_name: str, years: list) -> dict:
    """Apply accepted adjustments to P&L data and return adjusted copy."""
    import copy
    adjusted = copy.deepcopy(pnl_data)
    for adj in adjustments:
        if adj.get("action") == "reject":
            continue
        if adj.get("entity") != entity_name:
            continue
        line = adj.get("line_item", "")
        year = adj.get("year", "")
        amount = adj.get("adjustment_amount", 0) or 0
        if line in adjusted and year in adjusted[line]:
            orig = adjusted[line][year] or 0
            adjusted[line][year] = orig + amount
    return adjusted


def _get_adj_total(adjustments: list, entity_name: str, line_item: str, year: str) -> float:
    """Sum all accepted adjustments for a specific entity/line/year."""
    total = 0
    for adj in adjustments:
        if adj.get("action") == "reject":
            continue
        if adj.get("entity") == entity_name and adj.get("line_item") == line_item and adj.get("year") == year:
            total += adj.get("adjustment_amount", 0) or 0
    return total


# ── Main generator ───────────────────────────────────────────────────────────

def generate_consolidation_excel(
    extractions: list,
    adjustments: list,
    intercompany: list,
    entity_configs: list,
    fiscal_years: list,
    deal_code: str = "",
) -> bytes:
    """Generate the consolidated Excel workbook and return as bytes."""

    wb = Workbook()
    years = [str(y) for y in sorted(fiscal_years)]
    entities = [e["name"] for e in entity_configs]
    multi_entity = len(entities) > 1

    # ── Tab 1: Summary ───────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 30

    ws_summary.cell(row=1, column=1, value="MAX Solutions — Financial Consolidation").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value=f"Deal: {deal_code or 'N/A'}")
    ws_summary.cell(row=3, column=1, value=f"Entities: {', '.join(entities)}")
    ws_summary.cell(row=4, column=1, value=f"Generated: {date.today().isoformat()}")

    row = 6
    ws_summary.cell(row=row, column=1, value="Key Metrics").font = Font(bold=True, size=12)
    row += 1

    # Header
    ws_summary.cell(row=row, column=1, value="Metric")
    for ci, yr in enumerate(years):
        ws_summary.cell(row=row, column=2 + ci * 2, value=f"{yr} Reported")
        ws_summary.cell(row=row, column=3 + ci * 2, value=f"{yr} Normalized")
    _style_header_row(ws_summary, row, 1 + len(years) * 2)
    row += 1

    # Build consolidated reported and normalized P&L
    def _consolidated_pnl(kind="reported"):
        """Stack all entity P&Ls and return combined dict."""
        combined = {}
        for ext in extractions:
            ename = ext.get("entity_name", "")
            pnl = ext.get("pnl", {})
            if kind == "normalized":
                pnl = _apply_adjustments(pnl, adjustments, ename, years)
            for k, v in pnl.items():
                if k not in combined:
                    combined[k] = {yr: 0 for yr in years}
                for yr in years:
                    combined[k][yr] = (combined[k].get(yr) or 0) + ((v or {}).get(yr) or 0)
        return combined

    reported = _consolidated_pnl("reported")
    normalized = _consolidated_pnl("normalized")

    for metric_label, calc_key in [
        ("Revenue", "Total Revenue"),
        ("EBITDA", "EBITDA"),
        ("Net Profit", "Net Profit"),
    ]:
        ws_summary.cell(row=row, column=1, value=metric_label)
        for ci, yr in enumerate(years):
            _write_num(ws_summary, row, 2 + ci * 2, _calc_derived(calc_key, reported, yr))
            _write_num(ws_summary, row, 3 + ci * 2, _calc_derived(calc_key, normalized, yr))
        row += 1

    # Margins
    for label, num_key, denom_key in [
        ("EBITDA Margin", "EBITDA", "Total Revenue"),
    ]:
        ws_summary.cell(row=row, column=1, value=label)
        for ci, yr in enumerate(years):
            rev_r = _calc_derived(denom_key, reported, yr)
            rev_n = _calc_derived(denom_key, normalized, yr)
            ebitda_r = _calc_derived(num_key, reported, yr)
            ebitda_n = _calc_derived(num_key, normalized, yr)
            cell_r = ws_summary.cell(row=row, column=2 + ci * 2, value=ebitda_r / rev_r if rev_r else None)
            cell_n = ws_summary.cell(row=row, column=3 + ci * 2, value=ebitda_n / rev_n if rev_n else None)
            cell_r.number_format = _PCT_FMT
            cell_n.number_format = _PCT_FMT
        row += 1

    for col in range(2, 2 + len(years) * 2):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 2: Normalized P&L ────────────────────────────────────────────────
    ws_pnl = wb.create_sheet("Normalized P&L")
    ws_pnl.column_dimensions["A"].width = 30

    def _write_pnl_section(ws, start_row, entity_name, pnl_reported, pnl_adjusted):
        r = start_row
        ws.cell(row=r, column=1, value=entity_name).font = Font(bold=True, size=12)
        r += 1

        # Column headers: Line Item | Year1 Reported | Year1 Adj | Year1 Normalized | ...
        ws.cell(row=r, column=1, value="Line Item")
        col = 2
        for yr in years:
            ws.cell(row=r, column=col, value=f"{yr} Reported")
            ws.cell(row=r, column=col + 1, value=f"{yr} Adj")
            ws.cell(row=r, column=col + 2, value=f"{yr} Normalized")
            col += 3
        _style_header_row(ws, r, 1 + len(years) * 3)
        r += 1

        for key, label, is_total in _PNL_LINES:
            ws.cell(row=r, column=1, value=label)
            col = 2
            for yr in years:
                if is_total:
                    rep_val = _calc_derived(label, pnl_reported, yr)
                    adj_val = _calc_derived(label, pnl_adjusted, yr) - rep_val
                    norm_val = _calc_derived(label, pnl_adjusted, yr)
                else:
                    rep_val = (pnl_reported.get(key) or {}).get(yr)
                    norm_val = (pnl_adjusted.get(key) or {}).get(yr)
                    adj_val = ((norm_val or 0) - (rep_val or 0)) if rep_val is not None else None

                _write_num(ws, r, col, rep_val)
                _write_num(ws, r, col + 1, adj_val if adj_val else None)
                _write_num(ws, r, col + 2, norm_val)
                col += 3

            if is_total:
                _style_total_row(ws, r, 1 + len(years) * 3)
            r += 1

        return r + 1  # leave a gap

    current_row = 1
    for ext in extractions:
        ename = ext.get("entity_name", "")
        pnl_rep = ext.get("pnl", {})
        pnl_adj = _apply_adjustments(pnl_rep, adjustments, ename, years)
        current_row = _write_pnl_section(ws_pnl, current_row, ename, pnl_rep, pnl_adj)

    if multi_entity:
        current_row = _write_pnl_section(
            ws_pnl, current_row, "CONSOLIDATED",
            _consolidated_pnl("reported"), _consolidated_pnl("normalized"),
        )

    for col in range(2, 2 + len(years) * 3):
        ws_pnl.column_dimensions[get_column_letter(col)].width = 16

    # ── Tab 3: EBITDA Bridge ─────────────────────────────────────────────────
    ws_bridge = wb.create_sheet("EBITDA Bridge")
    ws_bridge.column_dimensions["A"].width = 30

    row = 1
    ws_bridge.cell(row=row, column=1, value="EBITDA Bridge — Reported to Normalized").font = Font(bold=True, size=12)
    row += 2

    ws_bridge.cell(row=row, column=1, value="Adjustment Category")
    for ci, yr in enumerate(years):
        ws_bridge.cell(row=row, column=2 + ci, value=yr)
    _style_header_row(ws_bridge, row, 1 + len(years))
    row += 1

    # Reported EBITDA
    ws_bridge.cell(row=row, column=1, value="Reported EBITDA")
    for ci, yr in enumerate(years):
        _write_num(ws_bridge, row, 2 + ci, _calc_derived("EBITDA", reported, yr))
    _style_subheader_row(ws_bridge, row, 1 + len(years))
    row += 1

    # Group adjustments by category
    categories = ["owner_expense", "related_party", "tax_adjustment", "non_recurring", "intercompany"]
    cat_labels = {
        "owner_expense": "Owner Personal Expenses",
        "related_party": "Related-Party Adjustments",
        "tax_adjustment": "Tax-Motivated Adjustments",
        "non_recurring": "Non-Recurring Items",
        "intercompany": "Intercompany Eliminations",
    }

    for cat in categories:
        cat_adjs = [a for a in adjustments if a.get("category") == cat and a.get("action") != "reject"]
        if not cat_adjs:
            continue
        ws_bridge.cell(row=row, column=1, value=cat_labels.get(cat, cat))
        for ci, yr in enumerate(years):
            total = sum(a.get("ebitda_impact", 0) or 0 for a in cat_adjs if a.get("year") == yr)
            _write_num(ws_bridge, row, 2 + ci, total if total else None)
        row += 1

    # Normalized EBITDA
    ws_bridge.cell(row=row, column=1, value="Normalized EBITDA")
    for ci, yr in enumerate(years):
        _write_num(ws_bridge, row, 2 + ci, _calc_derived("EBITDA", normalized, yr))
    _style_total_row(ws_bridge, row, 1 + len(years))
    row += 1

    # Delta
    row += 1
    ws_bridge.cell(row=row, column=1, value="EBITDA Uplift (%)").font = Font(bold=True)
    for ci, yr in enumerate(years):
        rep = _calc_derived("EBITDA", reported, yr)
        norm = _calc_derived("EBITDA", normalized, yr)
        if rep and rep != 0:
            cell = ws_bridge.cell(row=row, column=2 + ci, value=(norm - rep) / abs(rep))
            cell.number_format = _PCT_FMT

    for col in range(2, 2 + len(years)):
        ws_bridge.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 4: Balance Sheet ─────────────────────────────────────────────────
    has_bs = any(
        any((ext.get("balance_sheet", {}).get(k) or {}).get(yr) is not None
            for k, _ in _BS_LINES for yr in years)
        for ext in extractions
    )

    if has_bs:
        ws_bs = wb.create_sheet("Balance Sheet")
        ws_bs.column_dimensions["A"].width = 30

        row = 1
        for ext in extractions:
            ename = ext.get("entity_name", "")
            bs = ext.get("balance_sheet", {})

            ws_bs.cell(row=row, column=1, value=ename).font = Font(bold=True, size=12)
            row += 1

            ws_bs.cell(row=row, column=1, value="Line Item")
            for ci, yr in enumerate(years):
                ws_bs.cell(row=row, column=2 + ci, value=yr)
            _style_header_row(ws_bs, row, 1 + len(years))
            row += 1

            for key, label in _BS_LINES:
                ws_bs.cell(row=row, column=1, value=label)
                for ci, yr in enumerate(years):
                    val = (bs.get(key) or {}).get(yr)
                    _write_num(ws_bs, row, 2 + ci, val)
                row += 1
            row += 1

        for col in range(2, 2 + len(years)):
            ws_bs.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 5: Adjustments Detail ────────────────────────────────────────────
    ws_adj = wb.create_sheet("Adjustments Detail")

    adj_headers = ["Entity", "Year", "Line Item", "Original", "Adjustment", "Adjusted",
                   "Category", "Description", "Explanation", "Evidence", "Confidence", "Status"]
    for ci, h in enumerate(adj_headers):
        ws_adj.cell(row=1, column=ci + 1, value=h)
    _style_header_row(ws_adj, 1, len(adj_headers))

    for ri, adj in enumerate(adjustments, start=2):
        ws_adj.cell(row=ri, column=1, value=adj.get("entity", ""))
        ws_adj.cell(row=ri, column=2, value=adj.get("year", ""))
        ws_adj.cell(row=ri, column=3, value=adj.get("line_item", ""))
        _write_num(ws_adj, ri, 4, adj.get("original_amount"))
        _write_num(ws_adj, ri, 5, adj.get("adjustment_amount"))
        _write_num(ws_adj, ri, 6, adj.get("adjusted_amount"))
        ws_adj.cell(row=ri, column=7, value=adj.get("category", ""))
        ws_adj.cell(row=ri, column=8, value=adj.get("description", ""))
        ws_adj.cell(row=ri, column=9, value=adj.get("explanation", ""))
        ws_adj.cell(row=ri, column=10, value=adj.get("evidence_source", ""))
        ws_adj.cell(row=ri, column=11, value=adj.get("confidence", ""))
        ws_adj.cell(row=ri, column=12, value=adj.get("action", "accept").capitalize())

    col_widths = [18, 8, 25, 15, 15, 15, 18, 25, 40, 15, 12, 12]
    for ci, w in enumerate(col_widths):
        ws_adj.column_dimensions[get_column_letter(ci + 1)].width = w

    # ── Save to bytes ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
