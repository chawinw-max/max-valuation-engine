import math
import re
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string


def _normalize_ticker(ticker: str) -> str:
    """
    Strip LSEG sub-board suffixes so NTSCm.BK -> NTSC.BK, XOm.BK -> XO.BK.
    Also normalises KWNF.KL-style LSEG codes to uppercase.
    """
    if not ticker:
        return ticker
    m = re.match(r'^([A-Za-z0-9]+?)([a-z])(\.[A-Z]+)$', ticker)
    if m:
        return (m.group(1) + m.group(3)).upper()
    return ticker.upper()


_CORP_SUFFIX_RE = re.compile(
    r'\b(public\s+company\s+limited|berhad|bhd|pcl|co\.?,?\s*ltd\.?|limited|ltd\.?|inc\.?'
    r'|corp\.?|plc\.?|s\.?a\.?|n\.?v\.?|tbk\.?|pte\.?)\b',
    re.IGNORECASE,
)


def _normalize_company_name(name: str) -> str:
    """Strip legal suffixes and punctuation for fuzzy matching. 'Kawan Food Berhad' -> 'kawan food'"""
    n = _CORP_SUFFIX_RE.sub('', name.lower())
    return ' '.join(re.sub(r'[^a-z0-9\s]', ' ', n).split())


def _build_lseg_lookup(lseg_parsed_peers: list) -> dict:
    """
    Build a ticker -> peer dict with three match strategies:
    1. Exact uppercase ticker (e.g. SAUCE.BK, KWNF.KL)
    2. Normalised ticker stripping board suffix (NTSCm.BK -> NTSC.BK)
    3. Fuzzy company name after stripping legal suffixes (Kawan Food Berhad ~ Kawan Food Bhd)
    """
    lookup = {}
    for p in lseg_parsed_peers:
        original = p.get("identifier") or ""
        raw_upper = original.upper()
        if raw_upper:
            lookup[raw_upper] = p
        normalised = _normalize_ticker(original)  # normalize before uppercasing
        if normalised and normalised != raw_upper:
            lookup[normalised] = p
        cname = (p.get("company_name") or "").strip().lower()
        if cname:
            lookup[f"__name__{cname}"] = p
        norm_cname = _normalize_company_name(p.get("company_name") or "")
        if norm_cname and f"__name__{norm_cname}" not in lookup:
            lookup[f"__name__{norm_cname}"] = p
        # Also index by filename-derived ticker (e.g., "D.BK" from "D.BK.xlsx")
        fn_ticker = (p.get("filename_ticker") or "").upper()
        if fn_ticker and fn_ticker not in lookup:
            lookup[fn_ticker] = p
        # Index by filename as company name (e.g., "Wilmar International Ltd.xlsx")
        fn_raw = p.get("filename_raw") or ""
        if fn_raw and (' ' in fn_raw or len(fn_raw) > 10):
            fn_name_norm = _normalize_company_name(fn_raw)
            if fn_name_norm and f"__name__{fn_name_norm}" not in lookup:
                lookup[f"__name__{fn_name_norm}"] = p
    return lookup


def _fuzzy_name_match(lseg_by_ticker: dict, peer_company: str):
    """Last-resort: match if peer and LSEG share a distinctive word (3+ chars)."""
    peer_norm = _normalize_company_name(peer_company or "")
    if not peer_norm:
        return None
    peer_words = {w for w in peer_norm.split() if len(w) >= 3}
    best, best_score = None, 0
    for key, p in lseg_by_ticker.items():
        if not key.startswith("__name__"):
            continue
        lseg_words = {w for w in key[8:].split() if len(w) >= 3}
        overlap = peer_words & lseg_words
        if overlap:
            score = len(overlap) / min(len(peer_words), len(lseg_words))
            if score > best_score:
                best_score = score
                best = p
    return best if best_score > 0 else None


def _lseg_peer(lseg_by_ticker: dict, ticker: str, company_name: str) -> dict:
    """Look up LSEG data for a peer using ticker then company name fallbacks."""
    t = ticker or ""
    return (
        lseg_by_ticker.get(t.upper())
        or lseg_by_ticker.get(_normalize_ticker(t))
        or lseg_by_ticker.get(f"__name__{(company_name or '').strip().lower()}")
        or lseg_by_ticker.get(f"__name__{_normalize_company_name(company_name or '')}")
        or _fuzzy_name_match(lseg_by_ticker, company_name)
        or {}
    )


def _safe_write(sheet, cell_coord, value):
    """
    Writes a value to a cell ONLY if the cell does not already contain a formula.
    Handles merged cells by writing to the merge's anchor (top-left) cell.
    """
    cell = sheet[cell_coord]

    # If the target is a non-anchor merged cell, find and write to the anchor instead
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
        else:
            print(f"WARNING: {cell_coord} is a MergedCell with no resolvable anchor on {sheet.title}. Skipping.")
            return False

    if isinstance(cell.value, str) and cell.value.startswith('='):
        print(f"WARNING: Attempted to overwrite formula in {cell_coord} on sheet {sheet.title}. Skipping.")
        return False

    # Treat float NaN as empty — writing NaN to Excel causes #NUM! errors
    if isinstance(value, float) and math.isnan(value):
        return False

    cell.value = value
    return True


def _force_write(sheet, cell_coord, value):
    """
    Force-writes a value to a cell, even if it contains a formula.
    Used when the template structure changed and old formulas must be replaced.
    Handles merged cells like _safe_write.
    """
    cell = sheet[cell_coord]
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
        else:
            return False
    if isinstance(value, float) and math.isnan(value):
        return False
    cell.value = value
    return True


def _clear_cell(sheet, cell_coord):
    """Clear a cell's value (including formulas). For template migration."""
    cell = sheet[cell_coord]
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
        else:
            return
    cell.value = None


def _get_sheet(workbook, *names):
    """Return the first sheet matching any of the given names, or None.
    Handles the July 2026 template rename (numbered prefixes)."""
    for n in names:
        if n in workbook.sheetnames:
            return workbook[n]
    return None


def _detect_year_columns(sheet, header_row=3, min_col=2, max_col=14):
    """
    Map year (int) -> column letter by scanning a header row for year values.
    Robust to the template shifting year columns (e.g., P&L years moved from
    C-G to D-H when the FCF section was added).
    """
    mapping = {}
    for col in range(min_col, max_col + 1):
        v = sheet.cell(row=header_row, column=col).value
        try:
            y = int(float(v))
        except (TypeError, ValueError):
            continue
        if 2000 <= y <= 2100:
            mapping[y] = get_column_letter(col)
    return mapping


def inject_phase1_data(workbook, data, available_years):
    """
    Injects Phase 1 data (Preliminary P&L and Business Model) into the workbook.
    `data` is the JSON parsed dict from Gemini.
    """
    # 1. Preliminary P&L
    pnl = _get_sheet(workbook, '2. Preliminary P&L', 'Preliminary P&L')
    pnl_year_col = {}
    if pnl is not None:
        sheet = pnl

        # Year columns are detected from the row-3 header (D-H in the current
        # template — they shifted right when the FCF section was added).
        pnl_year_col = _detect_year_columns(sheet, header_row=3)
        if not pnl_year_col:
            pnl_year_col = {2021: 'D', 2022: 'E', 2023: 'F', 2024: 'G', 2025: 'H'}

        # Currency/source labels sit in the column left of the first year col,
        # rows 2 and 3 (C2/C3 in the current template).
        first_year_col_idx = min(
            column_index_from_string(c) for c in pnl_year_col.values()
        )
        label_col = get_column_letter(first_year_col_idx - 1)
        years_str = "-".join(map(str, available_years))
        _safe_write(sheet, f'{label_col}2', "Currency : Thai Baht")
        _safe_write(sheet, f'{label_col}3',
                    f"Source : {data.get('client_name', 'Company')} Financial Statements ({years_str})")

        financials = data.get('financials', {})

        # P&L row map (EBIT-first layout — rows unchanged across template revs):
        #   Row 6: Sales, Row 7: Other Rev, Row 8: Total Rev [FORMULA]
        #   Row 11: COGS, Row 14: GP [FORMULA]
        #   Row 18: Sales Exp, Row 19: Admin Exp, Row 20: Other Exp
        #   Row 24: EBIT [FORMULA], Row 27: D&A, Row 28: EBITDA [FORMULA]
        #   Row 32: Interest, Row 33: EBT [FORMULA]
        #   Row 36: Tax, Row 39: NP [FORMULA]
        row_map = {
            'sales_and_services': 6,
            'other_revenues': 7,
            'cost_of_goods_sold': 11,
            'sales_expenses': 18,
            'administrative_expenses': 19,
            'other_expenses': 20,
            'depreciation_amortization': 27,
            'interest_expenses': 32,
            'tax': 36,
        }

        for year in available_years:
            col_letter = pnl_year_col.get(year)
            if not col_letter:
                continue
            for key, row_num in row_map.items():
                val = financials.get(key, {}).get(str(year))
                if val is not None:
                    _safe_write(sheet, f'{col_letter}{row_num}', val)

    # 1b. Balance Sheet (new tab, July 2026) — from AUDITED financial statements
    bs_sheet = _get_sheet(workbook, '3. Balance Sheet', 'Balance Sheet')
    bs_year_col = {}
    if bs_sheet is not None:
        bs_year_col = _detect_year_columns(bs_sheet, header_row=3)
        if not bs_year_col:
            bs_year_col = {2021: 'C', 2022: 'D', 2023: 'E', 2024: 'F', 2025: 'G'}

        balance_sheet = data.get('balance_sheet', {}) or {}
        if balance_sheet:
            _safe_write(bs_sheet, 'B3', "Source : Audited Financial Statements")

            # Balance Sheet row map (inputs only — subtotals are formulas):
            #   Assets: 6 Cash, 7 AR, 8 ST loans receivable, 9 Inventories, 12 PPE net
            #   Liabilities: 18 AP, 19 ST loans, 20 Other current liab, 23 LT loans
            #   Equity: 29 Paid-up capital, 30 Retained earnings
            bs_row_map = {
                'cash_and_equivalents': 6,
                'accounts_receivable': 7,
                'short_term_loans_receivable': 8,
                'inventories': 9,
                'ppe_net': 12,
                'accounts_payable': 18,
                'short_term_loans': 19,
                'other_current_liabilities': 20,
                'long_term_loans': 23,
                'paid_up_capital': 29,
                'retained_earnings': 30,
            }
            for year in available_years:
                col_letter = bs_year_col.get(year)
                if not col_letter:
                    continue
                for key, row_num in bs_row_map.items():
                    val = balance_sheet.get(key, {}).get(str(year))
                    if val is not None:
                        _safe_write(bs_sheet, f'{col_letter}{row_num}', val)

    # 1c. P&L Unlevered FCF section (rows 42-60) — link working-capital and
    # PPE inputs to the Balance Sheet so the FCF waterfall computes.
    #   Row 51: Accounts receivable  ← BS row 7
    #   Row 52: Accounts payable     ← BS row 18
    #   Row 56: Ending PPE           ← BS row 12 (same year)
    #   Row 57: Beginning PPE        ← BS row 12 (prior year)
    if (pnl is not None and bs_sheet is not None
            and str(pnl['B42'].value or '').strip() == 'Free Cash Flow'):
        bs_title = bs_sheet.title
        for year in available_years:
            pcol = pnl_year_col.get(year)
            bcol = bs_year_col.get(year)
            if pcol and bcol:
                _safe_write(pnl, f'{pcol}51', f"='{bs_title}'!{bcol}7")
                _safe_write(pnl, f'{pcol}52', f"='{bs_title}'!{bcol}18")
                _safe_write(pnl, f'{pcol}56', f"='{bs_title}'!{bcol}12")
            prev_bcol = bs_year_col.get(year - 1)
            if pcol and prev_bcol and (year - 1) in available_years:
                _safe_write(pnl, f'{pcol}57', f"='{bs_title}'!{prev_bcol}12")

    # 2. Business Model (6-section layout: A through F)
    sheet = _get_sheet(workbook, '1. Business Model', 'Business Model')
    if sheet is not None:
        bm = data.get('business_model', {})
        client_name_bm = data.get('client_name', 'Company')

        # A1: Title
        _safe_write(sheet, 'A1', f"{client_name_bm.upper()} – BUSINESS MODEL OVERVIEW")

        # Section A: rows 4-13 (10 fields in B column)
        sec_a = bm.get('section_a', {}) or {}
        rows_a = ['company_name', 'founded', 'type', 'registered_capital',
                  'shareholders', 'location', 'focus', 'certifications',
                  'operating_history', 'website']
        for i, field in enumerate(rows_a):
            _safe_write(sheet, f'B{4+i}', sec_a.get(field))

        # Section B: rows 16-22 (7 fields in B column)
        sec_b = bm.get('section_b', {}) or {}
        rows_b = ['primary_revenue', 'key_clients', 'pricing_strategy',
                  'customer_mix', 'client_segmentation', 'credit_terms',
                  'secondary_revenue']
        for i, field in enumerate(rows_b):
            _safe_write(sheet, f'B{16+i}', sec_b.get(field))

        # Section C: rows 25-30 (moats, A=title, B=description)
        # Clear all 6 slots first — the template ships with sample moats and a
        # shorter moat list would otherwise leave stale sample rows behind.
        sec_c = bm.get('section_c', []) or []
        for i in range(6):
            _safe_write(sheet, f'A{25+i}', None)
            _safe_write(sheet, f'B{25+i}', None)
        for i, moat in enumerate(sec_c[:6]):
            _safe_write(sheet, f'A{25+i}', moat.get('title'))
            _safe_write(sheet, f'B{25+i}', moat.get('description'))

        # Section D: rows 33-40 (8 fields in B column)
        sec_d = bm.get('section_d', {}) or {}
        rows_d = ['revenue_fy2022', 'revenue_fy2023', 'revenue_fy2024',
                  'revenue_fy2025', 'gross_margin', 'ebitda_normalization',
                  'liabilities', 'cash_flow_note']
        for i, field in enumerate(rows_d):
            _safe_write(sheet, f'B{33+i}', sec_d.get(field))

        # Section E: rows 43-48 (6 fields in B column)
        sec_e = bm.get('section_e', {}) or {}
        rows_e = ['seller', 'exit_motivation', 'asking_price',
                  'property', 'transaction_type', 'next_steps']
        for i, field in enumerate(rows_e):
            _safe_write(sheet, f'B{43+i}', sec_e.get(field))

        # Section F: rows 51-57 (7 fields in B column)
        sec_f = bm.get('section_f', {}) or {}
        rows_f = ['total_staff', 'shift_structure', 'inbound_process',
                  'facility', 'machinery', 'logistics_fleet', 'it_systems']
        for i, field in enumerate(rows_f):
            _safe_write(sheet, f'B{51+i}', sec_f.get(field))

def inject_phase2_data(workbook, selected_peers, not_selected_peers, rejection_rationales):
    """
    Comps Profile tab (formerly 'Furniture'):
      Row 1 headers: A=Identifier, B=Company Name, C=TRBC Activity,
                     D=Country, E=Business Description, F=Market Cap
      Row 2:  'Selected' band + G2 'Reasoning'  → selected peers rows 3–9
      Row 10: 'Not Selected' band + G10 'Reasoning' → not-selected rows 11+
    """
    sheet = _get_sheet(workbook, 'Comps Profile', 'Furniture')
    if sheet is None:
        return

    # Selected peers — first 7 fit in rows 3–9
    for i, peer in enumerate(selected_peers[:7]):
        row = 3 + i
        _safe_write(sheet, f'A{row}', peer.get('identifier'))
        _safe_write(sheet, f'B{row}', peer.get('company_name'))
        _safe_write(sheet, f'C{row}', peer.get('trbc_activity'))
        _safe_write(sheet, f'D{row}', peer.get('country'))
        _safe_write(sheet, f'E{row}', peer.get('business_description'))
        _safe_write(sheet, f'F{row}', peer.get('market_cap_thb_m'))
        # Reasoning for selection: prefer ring_justification, otherwise blank
        _safe_write(sheet, f'G{row}', peer.get('ring_justification') or "Selected as comparable peer")

    # Not selected peers — start at row 11
    for i, peer in enumerate(not_selected_peers):
        row = 11 + i
        ticker = peer.get('identifier')
        _safe_write(sheet, f'A{row}', ticker)
        _safe_write(sheet, f'B{row}', peer.get('company_name'))
        _safe_write(sheet, f'C{row}', peer.get('trbc_activity'))
        _safe_write(sheet, f'D{row}', peer.get('country'))
        _safe_write(sheet, f'E{row}', peer.get('business_description'))
        _safe_write(sheet, f'F{row}', peer.get('market_cap_thb_m'))
        _safe_write(sheet, f'G{row}', rejection_rationales.get(ticker, "Not a strong fit"))


def inject_phase3_data(workbook, deep_dive, lseg_parsed_peers, selected_peers, deal_code, client_name, latest_year):
    """
    Comparison tab + Appendix Hist Trading Performan tab.
    Peer order is taken from `selected_peers` (Phase 2 ordering).
    """
    qualitative = {q.get('identifier'): q for q in deep_dive.get('qualitative', [])}
    financials = {f.get('identifier'): f for f in deep_dive.get('financials_comparison', [])}
    lseg_by_ticker = _build_lseg_lookup(lseg_parsed_peers)

    # ----- Comparison tab -----
    sheet = _get_sheet(workbook, '4. Comparison', '3. Comparison', 'Comparison')
    if sheet is not None:

        # Section I. Peers Profile — rows 5–10, columns C–J (max 6 peers)
        # Column K (Adjustment to Peer Multiple) is an XLOOKUP formula — don't touch.
        for i, peer in enumerate(selected_peers[:6]):
            row = 5 + i
            ticker = peer.get('identifier')
            q = qualitative.get(ticker, {})
            _safe_write(sheet, f'C{row}', ticker)
            _safe_write(sheet, f'D{row}', peer.get('company_name'))
            _safe_write(sheet, f'E{row}', q.get('core_business_model'))
            _safe_write(sheet, f'F{row}', q.get('product_focus'))
            _safe_write(sheet, f'G{row}', q.get('similarity'))
            _safe_write(sheet, f'H{row}', q.get('comparison_points'))
            _safe_write(sheet, f'I{row}', q.get('differentiation_points'))
            _safe_write(sheet, f'J{row}', peer.get('country'))

        # Section II. Peers Financial Profile — rows 18–23 (max 6 peers)
        # FIVE year blocks (Revenue col, EBITDA col, Margin col):
        #   2021: D/E/F, 2022: G/H/I, 2023: J/K/L, 2024: M/N/O, 2025: P/Q/R
        # C column has =C5..=C10 formulas linking to Section I tickers — skip C.
        year_blocks = [
            ('2021', 'D', 'E', 'F'), ('2022', 'G', 'H', 'I'),
            ('2023', 'J', 'K', 'L'), ('2024', 'M', 'N', 'O'),
            ('2025', 'P', 'Q', 'R'),
        ]

        # Clear all peer input cells first
        for i in range(6):
            row = 18 + i
            for _, rev_c, ebitda_c, margin_c in year_blocks:
                _force_write(sheet, f'{rev_c}{row}', None)
                _force_write(sheet, f'{ebitda_c}{row}', None)
                _force_write(sheet, f'{margin_c}{row}', None)

        for i, peer in enumerate(selected_peers[:6]):
            row = 18 + i
            ticker = peer.get('identifier')
            f = financials.get(ticker, {})
            for year, rev_c, ebitda_c, margin_c in year_blocks:
                rev = f.get(f'revenue_{year}')
                ebitda = f.get(f'ebitda_{year}')
                _force_write(sheet, f'{rev_c}{row}', rev)
                _force_write(sheet, f'{ebitda_c}{row}', ebitda)
                # Margin column feeds the Scale Discount premium (rows 33-36)
                if rev is not None and ebitda is not None:
                    _force_write(
                        sheet, f'{margin_c}{row}',
                        f'=IFERROR({ebitda_c}{row}/{rev_c}{row},"")'
                    )

    # ----- Summary tab — peer multiples table -----
    # C = ticker; E-I = EV/EBITDA 2021-2025; J-N = EV/Revenue; O-S = P/E.
    # These are INPUT cells feeding the Multiples Summary (TRANSPOSE array
    # formulas up top) and the Median/Average rows below the table.
    # The table has moved between template revisions (rows 29-34, then 33-38),
    # so locate it by finding the 'Ticker' header in column C.
    # D column (Country) is an XLOOKUP formula — don't touch.
    summary = _get_sheet(workbook, '0. Summary', 'Summary')
    if summary is not None:
        peer_start_row = None
        for r in range(20, 45):
            if str(summary.cell(row=r, column=3).value or '').strip() == 'Ticker':
                peer_start_row = r + 1
                break
        if peer_start_row is None:
            peer_start_row = 33  # current template default

        summary_years = ['2021', '2022', '2023', '2024', '2025']
        metric_start_cols = [
            ('ev_ebitda', 5),    # E-I  (col 5-9)
            ('ev_revenue', 10),  # J-N  (col 10-14)
            ('pe', 15),          # O-S  (col 15-19)
        ]
        for i, peer in enumerate(selected_peers[:6]):
            row = peer_start_row + i
            ticker = peer.get('identifier')
            _safe_write(summary, f'C{row}', ticker)
            lseg = _lseg_peer(lseg_by_ticker, ticker, peer.get('company_name'))
            for metric, start_col in metric_start_cols:
                values = lseg.get(metric, {}) or {}
                for y_idx, year in enumerate(summary_years):
                    v = values.get(year)
                    if v is not None:
                        col_letter = get_column_letter(start_col + y_idx)
                        _safe_write(summary, f'{col_letter}{row}', v)

    # ----- Appendix Hist Trading Performance tab -----
    sheet = _get_sheet(workbook, '5. Appendix Hist Trading Perfor',
                       '4. Appendix Hist Trading Perfor',
                       'Appendix Hist Trading Performan')
    if sheet is not None:

        # Header inputs (C4 is =TODAY() in the new template — don't touch)
        _safe_write(sheet, 'B2', f"Project {deal_code} - {client_name} Comps Tables")

        # Actual template structure (annual layout):
        #   EV/EBITDA: 6 peer rows at 9-14
        #     B=company name, C=ticker, D=2021, E=2022, F=2023, G=2024, H=2025
        #     I=Average formula, J=Median formula (don't touch)
        #     K-N=LTM quarterly (leave empty), O-P=LTM formulas (don't touch)
        #   P/E: 6 peer rows at 20-25
        #     C=formula linking to EV/EBITDA ticker (don't write to C)
        #     D-H=annual P/E data
        #   EV/Revenue: 6 peer rows at 31-36
        #     C=formula linking to EV/EBITDA ticker (don't write to C)
        #     D-H=annual EV/Revenue data

        # Annual data column mapping: 2021→D, 2022→E, 2023→F, 2024→G, 2025→H
        years_annual_cols = [
            ('2021', 'D'), ('2022', 'E'), ('2023', 'F'), ('2024', 'G'), ('2025', 'H'),
        ]

        # Clear INPUT cells for all 6 peer slots before writing
        for _r_off in range(6):
            # EV/EBITDA rows 9-14: clear B (col 2) and D-H (cols 4-8)
            for _col in [2, 4, 5, 6, 7, 8]:
                _cell = sheet.cell(row=9 + _r_off, column=_col)
                if not isinstance(_cell, MergedCell) and not (
                    isinstance(_cell.value, str) and _cell.value.startswith('=')
                ):
                    _cell.value = None
            # P/E rows 20-25: clear D-H (cols 4-8)
            for _col in [4, 5, 6, 7, 8]:
                _cell = sheet.cell(row=20 + _r_off, column=_col)
                if not isinstance(_cell, MergedCell) and not (
                    isinstance(_cell.value, str) and _cell.value.startswith('=')
                ):
                    _cell.value = None
            # EV/Revenue rows 31-36: clear D-H (cols 4-8)
            for _col in [4, 5, 6, 7, 8]:
                _cell = sheet.cell(row=31 + _r_off, column=_col)
                if not isinstance(_cell, MergedCell) and not (
                    isinstance(_cell.value, str) and _cell.value.startswith('=')
                ):
                    _cell.value = None

        for i, peer in enumerate(selected_peers[:6]):
            ticker = peer.get('identifier')
            lseg = _lseg_peer(lseg_by_ticker, ticker, peer.get('company_name'))

            # Section 1: EV/EBITDA — rows 9-14
            r1 = 9 + i
            _safe_write(sheet, f'B{r1}', peer.get('company_name'))  # Company name
            _safe_write(sheet, f'C{r1}', ticker)
            ev_ebitda = lseg.get('ev_ebitda', {})
            for year, col in years_annual_cols:
                v = ev_ebitda.get(year)
                if v is not None:
                    _safe_write(sheet, f'{col}{r1}', v)

            # Section 2: P/E — rows 20-25 (C column is FORMULA, skip)
            r2 = 20 + i
            pe = lseg.get('pe', {})
            for year, col in years_annual_cols:
                v = pe.get(year)
                if v is not None:
                    _safe_write(sheet, f'{col}{r2}', v)

            # Section 3: EV/Revenue — rows 31-36 (C column is FORMULA, skip)
            r3 = 31 + i
            ev_rev = lseg.get('ev_revenue', {})
            for year, col in years_annual_cols:
                v = ev_rev.get(year)
                if v is not None:
                    _safe_write(sheet, f'{col}{r3}', v)


def inject_phase35_data(workbook, transactions, deal_code):
    """
    Precedent Transactions tab (OLD column structure).
    Columns B-I: B=target, C=acquirer, D=date, E=region,
                 F=relevance, G=deal_value_usd_m, H=ev_ebitda, I=notes
    Default header at row 8, data starts at row 9.
    """
    sheet = _get_sheet(workbook, '6. Precedent Transactions', '5. Precedent Transactions',
                       'Precedent Transactions', 'Precedent Tx', 'Precedent_Transactions')
    if sheet is None:
        return

    # Locate header row by scanning rows 1-15 for "Target" + "Acquir" markers
    header_row = None
    for r in range(1, 16):
        row_vals = [str(sheet.cell(row=r, column=c).value or '').strip().lower()
                    for c in range(1, 12)]
        joined = ' | '.join(row_vals)
        if 'target' in joined and 'acquir' in joined:
            header_row = r
            break
    if header_row is None:
        header_row = 8  # template default

    # Update the Relevance header (template ships with a sample client name)
    _safe_write(sheet, f'F{header_row}', f"Relevance to {deal_code}")

    # Clear demo data rows (template ships with up to ~9 sample deals)
    for r in range(header_row + 1, header_row + 13):
        for c_letter in ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'):
            cell = sheet[f'{c_letter}{r}']
            v = cell.value
            if v is not None and not (isinstance(v, str) and v.startswith('=')):
                cell.value = None

    # Write up to 7 transactions (OLD column layout: B-I)
    for i, tx in enumerate(transactions[:7]):
        row = header_row + 1 + i

        target = tx.get('target') or tx.get('Target Full Name') or ''
        acquirer = tx.get('acquirer') or tx.get('Acquiror Full Name') or ''
        date = tx.get('date') or tx.get('Date Announced') or ''
        region = tx.get('region') or tx.get('Target Nation') or ''
        relevance = tx.get('relevance') or ''
        deal_value_usd = tx.get('deal_value_usd_m') or tx.get('Deal Value (USD, Millions)')
        ev_ebitda = tx.get('ev_ebitda') or tx.get('Ratio of Enterprise Value to EBITDA')
        notes = tx.get('notes_and_caveats') or ''

        _safe_write(sheet, f'B{row}', target)
        _safe_write(sheet, f'C{row}', acquirer)
        _safe_write(sheet, f'D{row}', date)
        _safe_write(sheet, f'E{row}', region)
        _safe_write(sheet, f'F{row}', relevance)
        _safe_write(sheet, f'G{row}', deal_value_usd)
        _safe_write(sheet, f'H{row}', ev_ebitda)
        _safe_write(sheet, f'I{row}', notes)


def inject_phase4_data(workbook, data, latest_year):
    """
    Summary tab — MINIMAL writes. The template handles almost everything
    via formulas that auto-calculate from the P&L and Appendix data.
    Only write the project title and latest year reference.
    """
    sheet = _get_sheet(workbook, '0. Summary', 'Summary')
    if sheet is None:
        return

    deal_code = data.get('deal_code', 'DF-XXX')
    client_name = data.get('client_name', 'Company')

    # B2: Project title ("DF[XXX] [Code Name] - Comps Tables" in the template)
    _safe_write(sheet, 'B2', f"{deal_code} {client_name} - Comps Tables")

    # II. Sensitivity Analysis — locate the 'Rationale' header, then populate
    # the four @EV(M) scenario rows below it. The @EV column is two columns
    # left of Rationale (G when Rationale is I); the multiple column between
    # them is formula-driven (=G/$D$5) — don't touch.
    # Scenario ladder: THB 100M steps centred near the discounted
    # EBITDA-multiple median implied EV (E14), matching K'Vipin's convention.
    rationale_cell = None
    for r in range(10, 30):
        for c in range(6, 12):
            if str(sheet.cell(row=r, column=c).value or '').strip() == 'Rationale':
                rationale_cell = (r, c)
                break
        if rationale_cell:
            break
    if rationale_cell:
        hdr_row, rat_col = rationale_cell
        ev_col = get_column_letter(rat_col - 2)
        rat_col_letter = get_column_letter(rat_col)
        seed_row = hdr_row + 1
        scenarios = [
            f'=IFERROR(ROUND($E$14,-2)-100,"")',
            f'=IFERROR({ev_col}{seed_row}+100,"")',
            f'=IFERROR({ev_col}{seed_row + 1}+100,"")',
            f'=IFERROR({ev_col}{seed_row + 2}+100,"")',
        ]
        rationales = [
            "Conservative — below the median implied EV from the discounted EBITDA multiple",
            "Base case — approximately the median implied EV (EBITDA multiple approach)",
            "Upside — modest premium for growth prospects or buyer synergies",
            "Ceiling — full strategic premium scenario",
        ]
        for i in range(4):
            _safe_write(sheet, f'{ev_col}{seed_row + i}', scenarios[i])
            _safe_write(sheet, f'{rat_col_letter}{seed_row + i}', rationales[i])
